"""
Financeiro NFS-e — MVP 2.0 Local
Base de dados : .xlsx local
Backup opcional: Google Sheets
Lançamento    : manual ou via upload de XML NFS-e (ABRASF 2.04)
"""

from __future__ import annotations
import json
from datetime import date, datetime
from pathlib import Path
from typing import Any

import pandas as pd
import plotly.express as px
import streamlit as st
import xmltodict
from openpyxl import load_workbook, Workbook

# ──────────────────────────────────────────────
# CAMINHOS E CONSTANTES
# ──────────────────────────────────────────────
BASE_DIR   = Path(__file__).parent
DB_PATH    = BASE_DIR / "data" / "financeiro.xlsx"
QUEUE_PATH = BASE_DIR / "data" / "sync_queue.json"
LOGO_PATH  = BASE_DIR / "logo.png"

SHEET_RECEBER = "BACKUP - Contas a Receber"
SHEET_PAGAR   = "BACKUP - Contas a Pagar"

# IDs das planilhas Google — deixe "" para desativar backup
GS_ID_RECEBER = "1FovYUynLFexkSRR81UQxAVq9nT3UpBndG_t7up0A_C0"
GS_ID_PAGAR   = "1ZRMVQvW_q95WtxQqwnZwPC7GcsR721pNP4wxkHxJVS4"

HEADERS = [
    "ID", "Tipo", "DataLancamento", "DataVencimento", "DataCompetencia",
    "NomeContraparte", "CNPJContraparte", "Descricao",
    "ValorBruto", "Deducoes", "ValorLiquido",
    "ISS", "ISSRetido", "Aliquota",
    "Status", "NumeroNF", "CodigoVerificacao", "Observacoes",
]

STATUS_OPTIONS = ["Pendente", "Pago", "Recebido", "Cancelado"]
NUM_COLS       = ["ValorBruto", "Deducoes", "ValorLiquido", "ISS", "Aliquota"]

# ──────────────────────────────────────────────
# CSS — PALETA DA EMPRESA
# ──────────────────────────────────────────────
CSS = """
<style>
    /* ── Paleta da empresa — compatível com tema claro e escuro ── */

    /* Métricas */
    [data-testid="stMetric"] {
        border-left: 4px solid #48bbed;
        border-radius: 8px;
        padding: 12px 16px;
        box-shadow: 0 1px 4px rgba(0,0,0,0.10);
    }

    /* Botão primário / submit */
    .stButton > button[kind="primary"],
    .stFormSubmitButton > button {
        background: linear-gradient(90deg, #48bbed, #14d5ee) !important;
        color: #fff !important;
        border: none !important;
        border-radius: 6px !important;
        font-weight: 600 !important;
    }
    .stButton > button[kind="primary"]:hover,
    .stFormSubmitButton > button:hover { opacity: 0.88 !important; }

    /* Botões secundários */
    .stButton > button {
        border: 1px solid #48bbed !important;
        color: #48bbed !important;
        border-radius: 6px !important;
    }

    /* Abas internas */
    .stTabs [data-baseweb="tab-list"] {
        border-bottom: 2px solid #48bbed;
    }
    .stTabs [data-baseweb="tab"][aria-selected="true"] {
        color: #48bbed !important;
        border-bottom: 3px solid #14d5ee !important;
        font-weight: 600;
    }

    /* Divisor */
    hr { border-color: #48bbed44 !important; }

    /* Inputs com foco */
    input:focus, textarea:focus {
        border-color: #14d5ee !important;
        box-shadow: 0 0 0 2px #14d5ee33 !important;
    }

    /* Sidebar — gradiente da empresa sobre qualquer tema */
    [data-testid="stSidebar"] {
        background: linear-gradient(180deg, #0d1b2a 0%, #1a2e44 100%) !important;
    }
    [data-testid="stSidebar"] * {
        color: #e8f4fc !important;
    }
    [data-testid="stSidebar"] .stRadio label {
        font-size: 0.95rem;
        padding: 6px 0;
    }
</style>
"""


# ──────────────────────────────────────────────
# PARSE XML NFS-e (ABRASF 2.04)
# ──────────────────────────────────────────────
def _strip_ns(d: Any) -> Any:
    if isinstance(d, dict):
        return {k.split(":")[-1]: _strip_ns(v) for k, v in d.items()}
    if isinstance(d, list):
        return [_strip_ns(i) for i in d]
    return d


def _g(node: Any, *keys: str, default: str = "") -> str:
    cur: Any = node
    for k in keys:
        if not isinstance(cur, dict):
            return default
        cur = cur.get(k)
        if cur is None:
            return default
    return str(cur).strip() if cur not in ({}, [], "") else default


def parse_nfse_xml(xml_bytes: bytes) -> list[dict[str, str]]:
    try:
        raw = _strip_ns(xmltodict.parse(xml_bytes, process_namespaces=False))
    except Exception as e:
        raise ValueError(f"XML inválido: {e}") from e

    root  = next(iter(raw.values()))
    lista = root.get("ListaNfse", root)
    comps = lista.get("CompNfse", [])
    if isinstance(comps, dict):
        comps = [comps]

    results = []
    for comp in comps:
        try:
            results.append(_extract(comp))
        except Exception as e:
            st.warning(f"Nota ignorada na extração: {e}")
    return results


def _extract(comp: dict) -> dict[str, str]:
    nfse     = comp.get("Nfse", comp)
    inf      = nfse.get("InfNfse", nfse)
    val_nfse = inf.get("ValoresNfse", {}) or {}
    decl     = inf.get("DeclaracaoPrestacaoServico", {}) or {}
    inf_decl = decl.get("InfDeclaracaoPrestacaoServico", decl)
    servico  = inf_decl.get("Servico", {}) or {}
    val_serv = servico.get("Valores", {}) or {}
    prest_d  = inf_decl.get("Prestador", {}) or {}
    prest_sv = inf.get("PrestadorServico", {}) or {}
    tomador  = inf_decl.get("TomadorServico", {}) or {}
    tom_id   = tomador.get("IdentificacaoTomador", {}) or {}
    tom_cnpj = tom_id.get("CpfCnpj", {}) or {}
    pr_cnpj  = prest_d.get("CpfCnpj", {}) or {}

    dt_raw   = _g(inf, "DataEmissao")
    dt_clean = dt_raw[:10] if dt_raw else str(date.today())

    return {
        "NumeroNF":          _g(inf, "Numero"),
        "CodigoVerificacao": _g(inf, "CodigoVerificacao"),
        "DataLancamento":    dt_clean,
        "DataVencimento":    dt_clean,
        "DataCompetencia":   _g(inf_decl, "Competencia")[:10] if _g(inf_decl, "Competencia") else dt_clean,
        "PrestadorCNPJ":     _g(pr_cnpj, "Cnpj") or _g(pr_cnpj, "Cpf"),
        "PrestadorNome":     _g(prest_sv, "RazaoSocial") or _g(prest_sv, "NomeFantasia"),
        "TomadorCNPJ":       _g(tom_cnpj, "Cnpj") or _g(tom_cnpj, "Cpf"),
        "TomadorNome":       _g(tomador, "RazaoSocial"),
        "ValorBruto":        _g(val_serv, "ValorServicos"),
        "ValorDeducoes":     _g(val_serv, "ValorDeducoes"),
        "BaseCalculo":       _g(val_nfse, "BaseCalculo"),
        "Aliquota":          _g(val_nfse, "Aliquota"),
        "ValorISS":          _g(val_nfse, "ValorIss"),
        "ISSRetido":         "Sim" if _g(servico, "IssRetido") == "1" else "Não",
        "ValorLiquido":      _g(val_nfse, "ValorLiquidoNfse"),
        "Discriminacao":     _g(servico, "Discriminacao"),
        "InformacoesComp":   _g(inf_decl, "InformacoesComplementares"),
    }


# ──────────────────────────────────────────────
# BANCO DE DADOS LOCAL (.xlsx)
# ──────────────────────────────────────────────
def _ensure_db() -> None:
    DB_PATH.parent.mkdir(parents=True, exist_ok=True)
    if DB_PATH.exists():
        return
    wb = Workbook()
    wb.remove(wb.active)
    for sheet in [SHEET_RECEBER, SHEET_PAGAR]:
        ws = wb.create_sheet(sheet)
        ws.append(HEADERS)
    wb.save(DB_PATH)


def _safe_save(wb, path: Path) -> None:
    """Salva o workbook. Se o arquivo estiver aberto no Excel, exibe aviso claro."""
    try:
        wb.save(path)
    except PermissionError:
        st.error(
            "⚠️ **Não foi possível salvar.** O arquivo `financeiro.xlsx` está aberto no Excel. "
            "Feche-o e tente novamente."
        )
        st.stop()


@st.cache_data(ttl=5, show_spinner=False)
def load_sheet(sheet_name: str) -> pd.DataFrame:
    _ensure_db()
    try:
        df = pd.read_excel(DB_PATH, sheet_name=sheet_name, dtype=str).fillna("")
        for col in HEADERS:
            if col not in df.columns:
                df[col] = ""
        return df[HEADERS]
    except Exception as e:
        st.error(f"Erro ao ler planilha: {e}")
        return pd.DataFrame(columns=HEADERS)


def save_row(row: dict, sheet_name: str) -> None:
    _ensure_db()
    wb = load_workbook(DB_PATH)
    ws = wb[sheet_name]
    ws.append([row.get(h, "") for h in HEADERS])
    _safe_save(wb, DB_PATH)
    load_sheet.clear()


def update_cell(sheet_name: str, row_id: str, col: str, value: str) -> None:
    update_cells_bulk(sheet_name, [row_id], col, value)


def update_cells_bulk(sheet_name: str, row_ids: list[str], col: str, value: str) -> None:
    """Atualiza uma coluna para múltiplos IDs em uma única gravação."""
    _ensure_db()
    wb = load_workbook(DB_PATH)
    ws = wb[sheet_name]
    headers = [c.value for c in ws[1]]
    id_col  = headers.index("ID") + 1
    tgt_col = headers.index(col) + 1
    ids_set = set(row_ids)
    for row in ws.iter_rows(min_row=2):
        if str(row[id_col - 1].value) in ids_set:
            row[tgt_col - 1].value = value
    _safe_save(wb, DB_PATH)
    load_sheet.clear()


def delete_row_by_id(sheet_name: str, row_id: str) -> None:
    _ensure_db()
    wb = load_workbook(DB_PATH)
    ws = wb[sheet_name]
    headers = [c.value for c in ws[1]]
    id_col  = headers.index("ID") + 1
    for i, row in enumerate(ws.iter_rows(min_row=2), start=2):
        if str(row[id_col - 1].value) == row_id:
            ws.delete_rows(i)
            break
    _safe_save(wb, DB_PATH)
    load_sheet.clear()


def _next_id(sheet_name: str) -> str:
    df     = load_sheet(sheet_name)
    prefix = "R" if sheet_name == SHEET_RECEBER else "P"
    if df.empty or df["ID"].eq("").all():
        return f"{prefix}0001"
    nums = pd.to_numeric(df["ID"].str[1:], errors="coerce").dropna()
    return f"{prefix}{int(nums.max()) + 1:04d}" if not nums.empty else f"{prefix}0001"


# ──────────────────────────────────────────────
# FILA DE SINCRONIZAÇÃO
# ──────────────────────────────────────────────
def _load_queue() -> list[dict]:
    if not QUEUE_PATH.exists():
        return []
    try:
        return json.loads(QUEUE_PATH.read_text(encoding="utf-8"))
    except Exception:
        return []


def _save_queue(queue: list[dict]) -> None:
    QUEUE_PATH.parent.mkdir(parents=True, exist_ok=True)
    QUEUE_PATH.write_text(json.dumps(queue, ensure_ascii=False, indent=2), encoding="utf-8")


def enqueue(row: dict, sheet_name: str) -> None:
    q = _load_queue()
    q.append({"sheet": sheet_name, "row": row, "ts": datetime.now().isoformat()})
    _save_queue(q)


def _try_sync_queue() -> tuple[int, int]:
    queue = _load_queue()
    if not queue:
        return 0, 0
    try:
        from sync import append_to_sheet
    except ImportError:
        return 0, len(queue)
    ok, remaining = 0, []
    for item in queue:
        try:
            gs_id = GS_ID_RECEBER if item["sheet"] == SHEET_RECEBER else GS_ID_PAGAR
            if gs_id:
                append_to_sheet(item["row"], gs_id, item["sheet"])
            ok += 1
        except Exception:
            remaining.append(item)
    _save_queue(remaining)
    return ok, len(remaining)


def _backup(row: dict, sheet_name: str) -> None:
    gs_id = GS_ID_RECEBER if sheet_name == SHEET_RECEBER else GS_ID_PAGAR
    if not gs_id:
        return
    try:
        from sync import append_to_sheet
        append_to_sheet(row, gs_id, sheet_name)
    except Exception:
        enqueue(row, sheet_name)
        st.toast("⚠️ Backup Sheets offline — lançamento salvo na fila.", icon="⚠️")


# ──────────────────────────────────────────────
# HELPERS UI
# ──────────────────────────────────────────────
def _num(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    for c in NUM_COLS:
        if c in df.columns:
            df[c] = pd.to_numeric(df[c].astype(str).str.replace(",", "."), errors="coerce").fillna(0.0)
    return df


def _fmt(v: float) -> str:
    return f"R$ {v:,.2f}"


def _parse_float(s: str) -> float:
    try:
        return float(str(s).replace(",", "."))
    except Exception:
        return 0.0


def _parse_date(s: str) -> date:
    for fmt in ("%Y-%m-%d", "%Y-%m-%dT%H:%M:%S", "%d/%m/%Y"):
        try:
            return datetime.strptime(s[:10], fmt[:10]).date()
        except Exception:
            continue
    return date.today()


# ──────────────────────────────────────────────
# SIDEBAR — LOGO
# ──────────────────────────────────────────────
def _sidebar_logo() -> None:
    if LOGO_PATH.exists():
        st.image(str(LOGO_PATH), use_container_width=True)
    else:
        st.markdown(
            "<div style='text-align:center;padding:8px 0'>"
            "<span style='font-size:2.4rem'>🧾</span>"
            "<p style='font-size:0.7rem;color:#48bbed;margin:2px 0'>logo.png não encontrada</p>"
            "</div>",
            unsafe_allow_html=True,
        )


# ──────────────────────────────────────────────
# FORMULÁRIO DE LANÇAMENTO
# ──────────────────────────────────────────────
def _form_lancamento(tipo: str, sheet_name: str) -> None:
    st.markdown("#### 📎 Importar via XML (opcional)")
    uploaded = st.file_uploader(
        "Upload do XML NFS-e para preencher automaticamente",
        type=["xml"], key=f"xml_{tipo}",
        help="Campos preenchidos automaticamente — editáveis antes de salvar.",
    )

    pre: dict[str, str] = {}
    batch: list[dict[str, str]] = []

    if uploaded:
        try:
            notas = parse_nfse_xml(uploaded.read())
            if notas:
                pre   = notas[0]
                batch = notas[1:]
                st.success(f"✅ {len(notas)} nota(s) lida(s). Revise e salve abaixo.")
                if batch:
                    st.info(f"ℹ️ Formulário mostra a 1ª nota. As outras {len(batch)} serão salvas automaticamente.")
        except ValueError as e:
            st.error(str(e))

    st.markdown("#### ✏️ Dados do Lançamento")

    nome_label = "Tomador (quem paga)" if tipo == "Receber" else "Prestador (quem cobra)"
    cnpj_label = "CNPJ do Tomador"     if tipo == "Receber" else "CNPJ do Prestador"
    nome_pre   = pre.get("TomadorNome",    "") if tipo == "Receber" else pre.get("PrestadorNome", "")
    cnpj_pre   = pre.get("TomadorCNPJ",    "") if tipo == "Receber" else pre.get("PrestadorCNPJ", "")

    with st.form(f"form_{tipo}", clear_on_submit=True):
        c1, c2, c3 = st.columns(3)
        nome      = c1.text_input(f"{nome_label} *", value=nome_pre)
        cnpj      = c2.text_input(cnpj_label,         value=cnpj_pre)
        descricao = c3.text_input("Discriminação",    value=pre.get("Discriminacao", ""))

        c4, c5, c6 = st.columns(3)
        dt_lanc = c4.date_input("Data do Lançamento", value=_parse_date(pre.get("DataLancamento", str(date.today()))))
        dt_venc = c5.date_input("Data de Vencimento",  value=_parse_date(pre.get("DataVencimento",  str(date.today()))))
        dt_comp = c6.date_input("Competência",          value=_parse_date(pre.get("DataCompetencia", str(date.today()))))

        c7, c8, c9 = st.columns(3)
        val_bruto = c7.number_input("Valor Bruto (R$) *",  min_value=0.0, value=_parse_float(pre.get("ValorBruto",    "0")), format="%.2f")
        deducoes  = c8.number_input("Deduções (R$)",        min_value=0.0, value=_parse_float(pre.get("ValorDeducoes", "0")), format="%.2f")
        aliquota  = c9.number_input("Alíquota ISS (%)",     min_value=0.0, max_value=10.0,
                                    value=_parse_float(pre.get("Aliquota", "5")), format="%.2f")

        c10, c11, c12 = st.columns(3)
        base_xml  = _parse_float(pre.get("BaseCalculo", "0"))
        bruto_ded = _parse_float(pre.get("ValorBruto",   "0")) - _parse_float(pre.get("ValorDeducoes", "0"))
        red_pre   = round((1 - base_xml / bruto_ded) * 100, 2) if bruto_ded > 0 and base_xml > 0 else 0.0
        reducao   = c10.number_input("Redução Base ISS (%)", min_value=0.0, max_value=100.0, value=red_pre, format="%.2f")
        iss_ret   = c11.selectbox("ISS Retido?", ["Sim", "Não"],
                                  index=0 if pre.get("ISSRetido", "Não") == "Sim" else 1)
        status    = c12.selectbox("Status", STATUS_OPTIONS)

        c13, c14  = st.columns(2)
        num_nf    = c13.text_input("Número da NF",          value=pre.get("NumeroNF", ""))
        cod_verif = c14.text_input("Código de Verificação", value=pre.get("CodigoVerificacao", ""))
        obs       = st.text_area("Observações / Inf. Complementares",
                                 value=pre.get("InformacoesComp", ""), height=68)

        submitted = st.form_submit_button("💾 Salvar Lançamento", use_container_width=True, type="primary")

    if submitted:
        if not nome or val_bruto <= 0:
            st.error("Preencha ao menos Nome/Razão Social e Valor Bruto.")
            return

        def _build(nome_, cnpj_, desc_, dtl_, dtv_, dtc_,
                   vb_, ded_, aliq_, red_, issr_, stat_, nnf_, cv_, ob_) -> dict:
            base_  = (vb_ - ded_) * (1 - red_ / 100)
            v_iss_ = base_ * (aliq_ / 100)
            v_liq_ = vb_ - ded_ - (v_iss_ if issr_ == "Sim" else 0)
            return {
                "ID":                _next_id(sheet_name),
                "Tipo":              tipo,
                "DataLancamento":    str(dtl_),
                "DataVencimento":    str(dtv_),
                "DataCompetencia":   str(dtc_),
                "NomeContraparte":   nome_,
                "CNPJContraparte":   cnpj_,
                "Descricao":         desc_,
                "ValorBruto":        f"{vb_:.2f}",
                "Deducoes":          f"{ded_:.2f}",
                "ValorLiquido":      f"{v_liq_:.2f}",
                "ISS":               f"{v_iss_:.2f}",
                "ISSRetido":         issr_,
                "Aliquota":          f"{aliq_:.2f}",
                "Status":            stat_,
                "NumeroNF":          nnf_,
                "CodigoVerificacao": cv_,
                "Observacoes":       ob_,
            }

        row = _build(nome, cnpj, descricao, dt_lanc, dt_venc, dt_comp,
                     val_bruto, deducoes, aliquota, reducao, iss_ret,
                     status, num_nf, cod_verif, obs)
        try:
            save_row(row, sheet_name)
            _backup(row, sheet_name)
        except Exception as e:
            st.error(f"Erro ao salvar: {e}")
            return

        saved_extra = 0
        for extra in batch:
            ne = extra.get("TomadorNome",  "") if tipo == "Receber" else extra.get("PrestadorNome", "")
            ce = extra.get("TomadorCNPJ",  "") if tipo == "Receber" else extra.get("PrestadorCNPJ", "")
            be = _parse_float(extra.get("BaseCalculo", "0"))
            bd = _parse_float(extra.get("ValorBruto",  "0")) - _parse_float(extra.get("ValorDeducoes", "0"))
            re = round((1 - be / bd) * 100, 2) if bd > 0 and be > 0 else 0.0
            row_e = _build(
                ne, ce, extra.get("Discriminacao", ""),
                _parse_date(extra.get("DataLancamento", str(date.today()))),
                _parse_date(extra.get("DataVencimento",  str(date.today()))),
                _parse_date(extra.get("DataCompetencia", str(date.today()))),
                _parse_float(extra.get("ValorBruto",    "0")),
                _parse_float(extra.get("ValorDeducoes", "0")),
                _parse_float(extra.get("Aliquota", "5")),
                re,
                "Sim" if extra.get("ISSRetido") == "Sim" else "Não",
                "Pendente",
                extra.get("NumeroNF", ""), extra.get("CodigoVerificacao", ""),
                extra.get("InformacoesComp", ""),
            )
            try:
                save_row(row_e, sheet_name)
                _backup(row_e, sheet_name)
                saved_extra += 1
            except Exception:
                pass

        st.success(f"✅ {1 + saved_extra} lançamento(s) salvos! Último ID: **{row['ID']}**")


# ──────────────────────────────────────────────
# TABELA DE LANÇAMENTOS
# ──────────────────────────────────────────────
def _tabela(df: pd.DataFrame, sheet_name: str) -> None:
    if df.empty:
        st.info("Nenhum lançamento encontrado.")
        return

    df = _num(df)

    f1, f2, f3 = st.columns(3)
    sf = f1.multiselect("Status", STATUS_OPTIONS, default=STATUS_OPTIONS, key=f"sf_{sheet_name}")
    nf = f2.text_input("Buscar nome / CNPJ", key=f"nf_{sheet_name}")
    mf = f3.text_input("Mês (YYYY-MM)", placeholder="2026-02", key=f"mf_{sheet_name}")

    mask = df["Status"].isin(sf)
    if nf:
        mask &= (df["NomeContraparte"].str.contains(nf, case=False, na=False) |
                 df["CNPJContraparte"].str.contains(nf, case=False, na=False))
    if mf:
        mask &= df["DataVencimento"].str.startswith(mf)

    df_view = df[mask]
    st.dataframe(df_view, use_container_width=True, hide_index=True)

    t1, t2, t3 = st.columns(3)
    t1.metric("Registros",           len(df_view))
    t2.metric("Valor Bruto Total",   _fmt(df_view["ValorBruto"].sum()))
    t3.metric("Valor Líquido Total", _fmt(df_view["ValorLiquido"].sum()))

    st.divider()
    st.markdown("**Ações em lançamentos:**")

    all_ids = df_view["ID"].tolist()

    # ── Atualização em lote ──
    with st.expander("🔄 Atualizar Status (individual ou em lote)", expanded=True):
        m1, m2 = st.columns([3, 1])
        modo = m1.radio("Modo", ["Individual", "Intervalo de IDs", "Selecionar múltiplos"],
                        horizontal=True, key=f"modo_{sheet_name}")
        n_stat = m2.selectbox("Novo Status", STATUS_OPTIONS, key=f"ast_{sheet_name}")

        ids_alvo: list[str] = []

        if modo == "Individual":
            sel_id = st.selectbox("ID", ["—"] + all_ids, key=f"aid_{sheet_name}")
            if sel_id != "—":
                ids_alvo = [sel_id]

        elif modo == "Intervalo de IDs":
            prefix = "R" if sheet_name == SHEET_RECEBER else "P"
            nums   = sorted([int(i[1:]) for i in all_ids if i.startswith(prefix)])
            if nums:
                b1, b2 = st.columns(2)
                ini = b1.selectbox("De",  [f"{prefix}{n:04d}" for n in nums], key=f"ini_{sheet_name}")
                fim = b2.selectbox("Até", [f"{prefix}{n:04d}" for n in nums],
                                   index=len(nums)-1, key=f"fim_{sheet_name}")
                ini_num = int(ini[1:])
                fim_num = int(fim[1:])
                if ini_num <= fim_num:
                    ids_alvo = [f"{prefix}{n:04d}" for n in range(ini_num, fim_num + 1)
                                if f"{prefix}{n:04d}" in all_ids]
                    st.caption(f"{len(ids_alvo)} nota(s) no intervalo selecionado.")
                else:
                    st.warning("'De' deve ser menor ou igual a 'Até'.")

        else:  # Selecionar múltiplos
            ids_alvo = st.multiselect("Selecione os IDs", all_ids, key=f"multi_{sheet_name}")

        if st.button("✅ Aplicar Status", key=f"aupd_{sheet_name}",
                     type="primary", use_container_width=True):
            if ids_alvo:
                update_cells_bulk(sheet_name, ids_alvo, "Status", n_stat)
                st.success(f"{len(ids_alvo)} nota(s) → **{n_stat}**")
                st.rerun()
            else:
                st.warning("Nenhuma nota selecionada.")

    # ── Excluir individual ──
    with st.expander("🗑️ Excluir lançamento"):
        d1, d2 = st.columns([3, 1])
        del_id = d1.selectbox("ID para excluir", ["—"] + all_ids, key=f"adel_id_{sheet_name}")
        d2.write("")
        d2.write("")
        if d2.button("Excluir", key=f"adel_{sheet_name}", use_container_width=True):
            if del_id != "—":
                delete_row_by_id(sheet_name, del_id)
                st.success(f"{del_id} excluído.")
                st.rerun()

    st.download_button(
        "⬇️ Exportar CSV", df_view.to_csv(index=False).encode("utf-8-sig"),
        file_name=f"{sheet_name}.csv", mime="text/csv", key=f"dl_{sheet_name}",
    )


# ──────────────────────────────────────────────
# PÁGINAS
# ──────────────────────────────────────────────
def page_lancamentos(tipo: str) -> None:
    sheet_name = SHEET_RECEBER if tipo == "Receber" else SHEET_PAGAR
    icon = "📥" if tipo == "Receber" else "📤"
    st.header(f"{icon} Contas a {tipo}")
    tab_novo, tab_lista = st.tabs(["➕ Novo Lançamento", "📋 Lançamentos"])
    with tab_novo:
        _form_lancamento(tipo, sheet_name)
    with tab_lista:
        col_r, _ = st.columns([1, 5])
        with col_r:
            if st.button("🔄 Atualizar", key=f"rel_{tipo}"):
                load_sheet.clear()
        _tabela(load_sheet(sheet_name), sheet_name)


def page_dashboard() -> None:
    st.header("📊 Dashboard Financeiro")
    col_r, _ = st.columns([1, 7])
    with col_r:
        if st.button("🔄 Atualizar"):
            load_sheet.clear()

    df_cr = _num(load_sheet(SHEET_RECEBER))
    df_cp = _num(load_sheet(SHEET_PAGAR))

    receita = df_cr["ValorLiquido"].sum()
    custo   = df_cp["ValorLiquido"].sum()
    saldo   = receita - custo
    iss_tot = df_cr["ISS"].sum() + df_cp["ISS"].sum()
    pend_cr = df_cr[df_cr["Status"] == "Pendente"]["ValorLiquido"].sum() if not df_cr.empty else 0.0
    pend_cp = df_cp[df_cp["Status"] == "Pendente"]["ValorLiquido"].sum() if not df_cp.empty else 0.0

    k = st.columns(6)
    k[0].metric("Receita Líquida",      _fmt(receita))
    k[1].metric("Custo Total",          _fmt(custo))
    k[2].metric("Saldo",                _fmt(saldo), delta=f"{saldo:+,.2f}")
    k[3].metric("ISS Total",            _fmt(iss_tot))
    k[4].metric("A Receber (Pendente)", _fmt(pend_cr))
    k[5].metric("A Pagar (Pendente)",   _fmt(pend_cp))

    st.divider()

    col1, col2 = st.columns(2)
    with col1:
        st.markdown("#### Receber por Contraparte (Top 10)")
        if not df_cr.empty and df_cr["NomeContraparte"].ne("").any():
            agg = df_cr.groupby("NomeContraparte", as_index=False)["ValorLiquido"].sum().nlargest(10, "ValorLiquido")
            fig = px.bar(agg, x="ValorLiquido", y="NomeContraparte", orientation="h",
                         text_auto=".2s", color="ValorLiquido", color_continuous_scale="Blues",
                         labels={"ValorLiquido": "R$", "NomeContraparte": ""})
            fig.update_layout(showlegend=False, yaxis={"categoryorder": "total ascending"},
                              margin=dict(l=0, r=0, t=4, b=0))
            st.plotly_chart(fig, use_container_width=True)
        else:
            st.info("Sem dados.")

    with col2:
        st.markdown("#### Pagar por Contraparte (Top 10)")
        if not df_cp.empty and df_cp["NomeContraparte"].ne("").any():
            agg2 = df_cp.groupby("NomeContraparte", as_index=False)["ValorLiquido"].sum().nlargest(10, "ValorLiquido")
            fig2 = px.bar(agg2, x="ValorLiquido", y="NomeContraparte", orientation="h",
                          text_auto=".2s", color="ValorLiquido", color_continuous_scale="Reds",
                          labels={"ValorLiquido": "R$", "NomeContraparte": ""})
            fig2.update_layout(showlegend=False, yaxis={"categoryorder": "total ascending"},
                               margin=dict(l=0, r=0, t=4, b=0))
            st.plotly_chart(fig2, use_container_width=True)
        else:
            st.info("Sem dados.")

    st.markdown("#### Fluxo Mensal")
    frames = []
    for df_t, lbl in [(df_cr, "Receber"), (df_cp, "Pagar")]:
        if df_t.empty:
            continue
        tmp = df_t.copy()
        tmp["Mes"] = pd.to_datetime(tmp["DataVencimento"], errors="coerce").dt.to_period("M").astype(str)
        agg_t = tmp.dropna(subset=["Mes"]).groupby("Mes", as_index=False)["ValorLiquido"].sum()
        agg_t["Tipo"] = lbl
        frames.append(agg_t)
    if frames:
        fig3 = px.line(pd.concat(frames), x="Mes", y="ValorLiquido", color="Tipo", markers=True,
                       labels={"Mes": "Mês", "ValorLiquido": "R$", "Tipo": ""},
                       color_discrete_map={"Receber": "#48bbed", "Pagar": "#e05c5c"})
        st.plotly_chart(fig3, use_container_width=True)

    col3, col4 = st.columns(2)
    for df_t, lbl, col in [(df_cr, "Receber", col3), (df_cp, "Pagar", col4)]:
        with col:
            st.markdown(f"#### Status — {lbl}")
            if not df_t.empty:
                agg_s = df_t.groupby("Status", as_index=False)["ValorLiquido"].sum()
                fig_s = px.pie(agg_s, names="Status", values="ValorLiquido", hole=0.4,
                               color_discrete_sequence=["#48bbed", "#14d5ee", "#0d8fa8", "#a8d8ea"])
                fig_s.update_traces(textinfo="percent+label")
                st.plotly_chart(fig_s, use_container_width=True)


def page_calculadora() -> None:
    st.header("🧮 Calculadora de Comissão")
    st.caption("Modelo base de comissionamento para corretoras de seguros")

    with st.expander("ℹ️ Como funciona o modelo"):
        st.markdown("""
        1. A **seguradora** paga comissão bruta à **corretora master**.
        2. A master repassa fração ao **corretor parceiro**.
        3. Incidem **ISS** e **IRRF** (PF) sobre o repasse.
        4. Valor líquido = base do DAS / carnê-leão.
        """)

    c1, c2, c3 = st.columns(3)
    premio   = c1.number_input("Prêmio Líquido (R$)",     min_value=0.0, value=10000.0, step=100.0, format="%.2f")
    com_pct  = c2.number_input("Comissão Seguradora (%)", min_value=0.0, max_value=100.0, value=20.0, format="%.2f")
    rep_pct  = c3.number_input("Repasse ao Corretor (%)", min_value=0.0, max_value=100.0, value=50.0, format="%.2f")

    c4, c5, c6 = st.columns(3)
    iss_pct  = c4.number_input("ISS (%)",        min_value=0.0, max_value=10.0,  value=5.0, format="%.2f")
    irrf_pct = c5.number_input("IRRF — PF (%)",  min_value=0.0, max_value=27.5,  value=0.0, format="%.2f")
    iss_ret  = c6.checkbox("ISS retido na fonte?", value=True)

    c7, c8  = st.columns(2)
    qtd     = c7.number_input("Qtd. de Apólices",     min_value=1, value=1, step=1)
    reducao = c8.number_input("Redução Base ISS (%)", min_value=0.0, max_value=100.0, value=0.0, step=5.0, format="%.2f")

    com_bruta  = premio * (com_pct / 100)
    v_repasse  = com_bruta * (rep_pct / 100)
    base_iss   = v_repasse * (1 - reducao / 100)
    v_iss      = base_iss * (iss_pct / 100)
    v_irrf     = v_repasse * (irrf_pct / 100)
    deducoes   = (v_iss if iss_ret else 0) + v_irrf
    v_liq      = v_repasse - deducoes
    ret_master = com_bruta - v_repasse

    st.divider()
    st.subheader("Resultado por Apólice")
    r1, r2, r3, r4 = st.columns(4)
    r1.metric("Comissão Bruta",      _fmt(com_bruta))
    r2.metric("Repasse Bruto",       _fmt(v_repasse))
    r3.metric("Deduções (ISS+IRRF)", _fmt(deducoes))
    r4.metric("💰 Repasse Líquido",  _fmt(v_liq))

    st.subheader(f"Resultado — {qtd} Apólice(s)")
    p1, p2, p3, p4 = st.columns(4)
    p1.metric("Repasse Bruto Total",   _fmt(v_repasse * qtd))
    p2.metric("ISS Total",             _fmt(v_iss * qtd))
    p3.metric("Retenção da Master",    _fmt(ret_master * qtd))
    p4.metric("💰 Líquido do Período", _fmt(v_liq * qtd))

    partes = {"Líquido Corretor": v_liq, "ISS": v_iss, "IRRF": v_irrf, "Retenção Master": ret_master}
    df_p = pd.DataFrame({"Parcela": list(partes.keys()), "Valor": list(partes.values())})
    fig = px.pie(df_p[df_p["Valor"] > 0], names="Parcela", values="Valor", hole=0.4,
                 color_discrete_sequence=["#48bbed", "#14d5ee", "#0d8fa8", "#a8d8ea"])
    fig.update_traces(textinfo="percent+label")
    st.plotly_chart(fig, use_container_width=True)

    with st.expander("🔍 Memória de Cálculo"):
        for k, v in {
            "Prêmio Líquido":                    _fmt(premio),
            f"Comissão ({com_pct}%)":            _fmt(com_bruta),
            f"Repasse ({rep_pct}% da comissão)": _fmt(v_repasse),
            f"Base ISS (redução {reducao}%)":    _fmt(base_iss),
            f"ISS ({iss_pct}% s/ base)":         _fmt(v_iss),
            f"IRRF ({irrf_pct}%)":               _fmt(v_irrf),
            "Repasse Líquido":                   _fmt(v_liq),
        }.items():
            st.markdown(f"- **{k}:** {v}")


def page_configuracoes() -> None:
    st.header("⚙️ Configurações")

    st.subheader("🗄️ Banco de Dados Local")
    _ensure_db()
    db_size = DB_PATH.stat().st_size / 1024 if DB_PATH.exists() else 0
    c1, c2 = st.columns(2)
    c1.info(f"📁 `{DB_PATH}`")
    c2.info(f"💾 Tamanho: `{db_size:.1f} KB`")

    df_cr = load_sheet(SHEET_RECEBER)
    df_cp = load_sheet(SHEET_PAGAR)
    i1, i2 = st.columns(2)
    i1.metric("Lançamentos — Receber", len(df_cr))
    i2.metric("Lançamentos — Pagar",   len(df_cp))

    st.divider()
    st.subheader("🖼️ Logo")
    if LOGO_PATH.exists():
        st.success(f"✅ `logo.png` carregada de `{LOGO_PATH}`")
        st.image(str(LOGO_PATH), width=200)
    else:
        st.warning(f"⚠️ `logo.png` não encontrada. Coloque o arquivo em: `{BASE_DIR}`")

    st.divider()
    st.subheader("☁️ Backup Google Sheets")
    queue = _load_queue()
    if not GS_ID_RECEBER and not GS_ID_PAGAR:
        st.warning("Backup desativado. Preencha `GS_ID_RECEBER` / `GS_ID_PAGAR` no topo do arquivo.")
    else:
        if queue:
            st.warning(f"⏳ **{len(queue)} lançamento(s)** na fila.")
            if st.button("🔄 Sincronizar agora"):
                ok, fail = _try_sync_queue()
                if ok:   st.success(f"{ok} sincronizado(s).")
                if fail: st.error(f"{fail} falhou(aram).")
        else:
            st.success("✅ Tudo sincronizado.")

    # ── Diagnóstico de conexão ──
    st.divider()
    st.subheader("🔍 Diagnóstico de Conexão")
    if st.button("▶️ Testar conexão com Google Sheets agora"):
        # Passo 1 — secrets.toml
        with st.status("Verificando configuração...", expanded=True) as status:
            st.write("🔑 Lendo secrets.toml...")
            try:
                creds_dict = dict(st.secrets["gcp_service_account"])
                email = creds_dict.get("client_email", "não encontrado")
                st.write(f"✅ Credenciais lidas. Service Account: `{email}`")
            except Exception as e:
                st.write(f"❌ Falha ao ler secrets.toml: `{e}`")
                status.update(label="❌ Erro nas credenciais", state="error")
                st.stop()

            # Passo 2 — autenticação
            st.write("🔐 Autenticando com Google...")
            try:
                from google.oauth2.service_account import Credentials
                import gspread
                creds = Credentials.from_service_account_info(
                    creds_dict,
                    scopes=[
                        "https://www.googleapis.com/auth/spreadsheets",
                        "https://www.googleapis.com/auth/drive",
                    ],
                )
                client = gspread.authorize(creds)
                st.write("✅ Autenticação OK.")
            except Exception as e:
                st.write(f"❌ Falha na autenticação: `{e}`")
                status.update(label="❌ Erro na autenticação", state="error")
                st.stop()

            # Passo 3 — acessar planilha Receber
            if GS_ID_RECEBER:
                st.write(f"📄 Acessando planilha Receber (`{GS_ID_RECEBER[:20]}...`)...")
                try:
                    sh = client.open_by_key(GS_ID_RECEBER)
                    st.write(f"✅ Planilha Receber acessada: `{sh.title}`")
                except Exception as e:
                    st.write(f"❌ Falha ao acessar planilha Receber: `{e}`")
                    status.update(label="❌ Erro ao acessar planilha", state="error")
                    st.stop()

            # Passo 4 — acessar planilha Pagar
            if GS_ID_PAGAR:
                st.write(f"📄 Acessando planilha Pagar (`{GS_ID_PAGAR[:20]}...`)...")
                try:
                    sh2 = client.open_by_key(GS_ID_PAGAR)
                    st.write(f"✅ Planilha Pagar acessada: `{sh2.title}`")
                except Exception as e:
                    st.write(f"❌ Falha ao acessar planilha Pagar: `{e}`")
                    status.update(label="❌ Erro ao acessar planilha Pagar", state="error")
                    st.stop()

            status.update(label="✅ Tudo funcionando!", state="complete")

    st.divider()
    st.subheader("📤 Exportar Base")
    e1, e2 = st.columns(2)
    with e1:
        if not df_cr.empty:
            st.download_button("⬇️ Receber (.csv)", df_cr.to_csv(index=False).encode("utf-8-sig"),
                               "receber.csv", "text/csv")
    with e2:
        if not df_cp.empty:
            st.download_button("⬇️ Pagar (.csv)", df_cp.to_csv(index=False).encode("utf-8-sig"),
                               "pagar.csv", "text/csv")

    st.divider()
    st.subheader("⚠️ Zona de Perigo")
    with st.expander("Limpar dados"):
        st.error("Ações irreversíveis.")
        d1, d2 = st.columns(2)
        if d1.button("🗑️ Apagar todos os Receber"):
            try:
                wb = load_workbook(DB_PATH)
                ws = wb[SHEET_RECEBER]
                ws.delete_rows(2, ws.max_row)
                _safe_save(wb, DB_PATH)
                load_sheet.clear()
                st.success("Apagado.")
            except PermissionError as e:
                st.error(str(e))
        if d2.button("🗑️ Apagar todos os Pagar"):
            try:
                wb = load_workbook(DB_PATH)
                ws = wb[SHEET_PAGAR]
                ws.delete_rows(2, ws.max_row)
                _safe_save(wb, DB_PATH)
                load_sheet.clear()
                st.success("Apagado.")
            except PermissionError as e:
                st.error(str(e))


# ──────────────────────────────────────────────
# MAIN
# ──────────────────────────────────────────────
def main() -> None:
    st.set_page_config(page_title="Financeiro MVP 2.0", page_icon="🧾", layout="wide")
    _ensure_db()

    # Injeta CSS da paleta da empresa
    st.markdown(CSS, unsafe_allow_html=True)

    if _load_queue() and (GS_ID_RECEBER or GS_ID_PAGAR):
        ok, _ = _try_sync_queue()
        if ok:
            st.toast(f"☁️ {ok} lançamento(s) sincronizado(s).", icon="✅")

    with st.sidebar:
        _sidebar_logo()
        st.markdown(
            "<h2 style='text-align:center;color:#48bbed !important;"
            "font-size:1.1rem;margin:4px 0 12px 0'>Financeiro 2.0</h2>",
            unsafe_allow_html=True,
        )
        st.divider()
        page = st.radio(
            "menu",
            ["📊 Dashboard", "📥 Contas a Receber",
             "📤 Contas a Pagar", "🧮 Calculadora", "⚙️ Configurações"],
            label_visibility="collapsed",
        )
        st.divider()
        q = _load_queue()
        if q:
            st.warning(f"⏳ {len(q)} na fila de sync")
        elif GS_ID_RECEBER or GS_ID_PAGAR:
            st.success("☁️ Sheets sincronizado")
        else:
            st.info("☁️ Backup desativado")

    if page == "📊 Dashboard":
        page_dashboard()
    elif page == "📥 Contas a Receber":
        page_lancamentos("Receber")
    elif page == "📤 Contas a Pagar":
        page_lancamentos("Pagar")
    elif page == "🧮 Calculadora":
        page_calculadora()
    elif page == "⚙️ Configurações":
        page_configuracoes()


if __name__ == "__main__":
    main()